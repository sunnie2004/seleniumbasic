#############################################################################
#excel.py
#author: xiaoqing sunny, date:2018-10-10/17
#function: read and write the excel
#############################################################################
#coding = utf-8

import time
import os
from selenium import webdriver
import  xlrd
import  xlwt
import xlutils.copy
import openpyxl
from datetime import date,datetime
class ExcelUtil():
#################################
#function: read the excel, result:ok
################################

    def dict_data(self,excelPath,sheetName):
        self.data = xlrd.open_workbook(excelPath)
        self.table = self.data.sheet_by_name(sheetName)
        self.count = len(self.data.sheets())
        self.rowNum = self.table.nrows
        self.colNum = self.table.ncols
        print("总页数:%s"%self.count)
        print("行数为:%s"%self.rowNum)
        print("列数为:%s"%self.colNum)

        if self.rowNum <= 1:
            print("总行小于1")
        else:
           for i in range(self.rowNum):
               values = self.table.row_values(i)
               for j in values:
                    print(j,end=" ")
               print("")
####################################

####################################
#append into the list, result:ok
####################################
    def dict_data1(self,excelPath,sheetName):
        self.data = xlrd.open_workbook(excelPath)
        self.table = self.data.sheet_by_name(sheetName)
        self.count = len(self.data.sheets())
        self.rowNum = self.table.nrows
        self.colNum = self.table.ncols
        self.keys = self.table.row_values(0)  #get the first row as the keys in dict
        print(self.keys)
        print("总页数:%s"%self.count)
        print("行数为:%s"%self.rowNum)
        print("列数为:%s"%self.colNum)

        if self.rowNum <= 1:
            print("总行小于1")
        else:                                    #return the list including the dict
            r = []                               #列表,装结果的序列
            i = 0
            x =0
            j = 1
            for i in range((self.rowNum)-1):    #遍历行的数据
                s = {}                          # 字典
                values = self.table.row_values(j)    #根据行号得到行的数据
                print(values)
                for x in range(self.colNum):
                    s[self.keys[x]] = values[x]
                j = j + 1
                r.append(s)                     #使用append添加列表元素

            return r

##########################################
#        else:                               # return the whole list   ok
#            r = []                          #列表,装结果的序列
#            i = 0
#            s = {}                            #字典
#            x =0
#            j = 0
#            for x in range(self.rowNum):      #遍历行的数据
#                values = self.table.row_values(x)    #根据行号得到行的数据
#                for i in range(self.colNum):
#                    s[j] = values[i]
#                    j = j + 1
#            r.append(s)                    #使用append添加列表元素
#            return r
#########################################


##########################################
#function: read the cell value about date, result:ok
#########################################
    def read_text(self,excelPath,sheetName):
         self.data = xlrd.open_workbook(excelPath)
         self.table = self.data.sheet_by_name(sheetName)
#        self.table = self.data.sheets()[0]
#        self.table = self.data.sheet_by_index(0)
         print(self.table.cell_value(0,5))
         print(self.table.row(0)[5].value)
         print(self.table.cell(0,5).ctype)
         print(self.table.cell(1,5).ctype)
         data_value = xlrd.xldate_as_tuple(self.table.cell_value(0,5),self.data.datemode)
         print(data_value)
         print(date(*data_value[:3]))
         print(date(*data_value[:3]).strftime("%Y/%m/%d"))
###############################################

##############################################
#read the merge cell value, result:ok
#(0,2,1,2) ---(row,row_range,col,col_range):row 0&1 merge, col 1 merge
#row_range and col_range not include in the merge
    def read_merge(self,excelPath,sheetName):
        self.data = xlrd.open_workbook(excelPath)
        self.table = self.data.sheet_by_name(sheetName)
        merge = []     #list
        print(self.table.merged_cells)
        print(self.table.cell_value(1,1))            #merged cell ""
        for(rlow,rhigh,clow,chigh) in self.table.merged_cells:
            merge.append([rlow,clow])
            print(rlow)
            print(clow)
            for index in merge:
                print(self.table.cell_value(index[0],index[1]))


##############################################################
#funciton: write the excel, result:  ok
    def write_msg(self,msg,sheet,outfile):
        if os.path.isfile(outfile):
            readbook = xlrd.open_workbook(outfile,formatting_info=True)
            workbook = xlutils.copy.copy(readbook)
            if workbook.get_sheet(sheet):
                sheet1 = workbook.get_sheet(sheet)
            else:
                sheet1 = workbook.add_sheet(sheet, cell_overwrite_ok=True)
#####################################################
#notice :    openpyxl does not support old.xls file
#            workbook = openpyxl.load_workbook(outfile)
#            sheet1 = workbook.creatsheet(sheet)
#####################################################
        else:
            workbook = xlwt.Workbook()  # 创建工作薄
            sheet1 = workbook.add_sheet(sheet, cell_overwrite_ok=True)
        col = 0
        for i in msg:
            sheet1.write(0, col, i)
            col = col + 1
        workbook.save(outfile)
#####################################################################

#####################################################################
#write data  from the txt.fle to excel ,result:ok
###################################################
    def write_text_to_Excel(self,sheet,inputfile,outfile):
        fopen = open(inputfile,'r')
        line = fopen.read()
        print(line)

        if os.path.exists(outfile):
            readbook = xlrd.open_workbook(outfile,formatting_info=True)
            workbook = xlutils.copy.copy(readbook)
#            workbook = openpyxl.load_workbook(outfile)
            if workbook.get_sheet(sheet):
                sheet1 = workbook.get_sheet(sheet)
            else:
                sheet1 = workbook.add_sheet(sheet, cell_overwrite_ok=True)
        else:
            workbook = xlwt.Workbook()  # 创建工作薄
            sheet1 = workbook.add_sheet(sheet, cell_overwrite_ok=True)
        arr_line = line[1:].split("*")  #find the rows
        print(arr_line)
        print(len(arr_line))
        for i in range(len(arr_line)):
            arr_cell = arr_line[i].split(" ")
            print(arr_cell)
            for j in range(len(arr_cell)):
                sheet1.write(i,j,arr_cell[j])
        workbook.save(outfile)
#######################################################################

if __name__ == "_main_":
    filepath = "D:\\test\\test02\\goods and customes.xlsx"
    sheetName = "testlogin"
    data = ExcelUtil()
#    data.read_merge(filepath,sheetName)
#    data.read_text(filepath,sheetName)
    print(data.dict_data1(filepath,sheetName))
#    data.dict_data(filepath,sheetName)

################################
#write one message to excel  ok
################################
#    msg1 = ["41861", "chris", "visa", "macy", "cosmetic"]
#    save_file = "D:\\test\\test02\\hello.xls"
#   sheetName1 = "goods"
#    data.write_msg(msg1,sheetName1,save_file)

############################################
#read the text file then write to excel ok
#    text_file = "D:\\test\\test02\\testing.txt"    #row starts by '*',col seperated by "tab"
#    sheetName2 = "test"
#    data.write_text_to_Excel(sheetName2,text_file,save_file)


